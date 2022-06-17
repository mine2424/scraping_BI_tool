from openpyxl import Workbook
from openpyxl import load_workbook


from constants.openpyxl_constants import OpenPyxlConstants


class OpenpyxlService:
    def __init__(self) -> None:
        self.sheet = None
        self.book = None
        self.excelFile = ''
        self.max_col = 1
        self.max_row = 2

    def add_max_col(self):
        self.max_col += 1

    def add_max_row(self):
        self.max_row += 1

    def clear_max_col_row(self):
        self.max_col = 0
        self.max_row = 0

    def init_openpyxl(self, fileName: str):
        initializedSheet = Workbook()
        self.excelFile = f'../excel_data/{fileName}.xlsx'
        initializedSheet.save(self.excelFile)
        self.book = load_workbook(self.excelFile)
        self.sheet = self.book.worksheets[0]
        return self.sheet

    def save_sheet(self, fileName: str = ''):
        if fileName == '':
            self.book.save(self.excelFile)
        else:
            self.book.save(filename=fileName)

    def create_title(self):
        for i, indexTitle in enumerate(OpenPyxlConstants.titles):
            self.sheet.cell(row=1, column=i+1).value = indexTitle
        self.save_sheet()

    def add_data_in_sheet(self, maker_name: str, vehicle_type_name: str, grade_name: str, spec_details: list):
        self.sheet.cell(row=self.max_row, column=1).value = maker_name
        self.sheet.cell(row=self.max_row, column=2).value = vehicle_type_name
        self.sheet.cell(row=self.max_row, column=3).value = grade_name

        title_column_index = 4
        # TODO: 電気とレギュラーで全然テーブル項目が違うのでtitleで紐付ける
        for i, spec_detail in enumerate(spec_details):
            # if OpenPyxlConstants.titlesで存在しないtitleだったら追加する
            title_column_index = i
            self.sheet.cell(
                row=self.max_row, column=i+4
            ).value = spec_detail.value
            if spec_detail.title not in OpenPyxlConstants.titles:
                self.sheet.cell(
                    row=1, column=title_column_index
                ).value = spec_detail.title
                title_column_index = i + 1
                pass

        title_column_index = 0

        self.add_max_row()
        self.save_sheet()

    def remove_space_col(self):
        # 特定の列列を削除する
        #　TODO: 後程、自動にするようにする
        self.sheet.delete_cols(5)
        self.save_sheet()
